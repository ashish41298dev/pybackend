"""
Admin reports — single source of truth for the 5 admin reports.

Each report function returns:
    (columns: list[(key, label)], rows: list[dict], meta: dict)

The shared `render_report` helper turns that into JSON, PDF, or XLSX.
"""
from __future__ import annotations

import io
import re
from datetime import datetime, timezone, timedelta
from typing import Any, Callable, Dict, List, Optional, Tuple

from fastapi import HTTPException
from fastapi.responses import StreamingResponse, JSONResponse


# ---------------------------------------------------------------------------
# Filter helpers
# ---------------------------------------------------------------------------
def _parse_dt(value: Optional[str], end_of_day: bool = False) -> Optional[datetime]:
    if not value:
        return None
    try:
        if "T" in value:
            dt = datetime.fromisoformat(value)
        else:
            dt = datetime.fromisoformat(value + ("T23:59:59" if end_of_day else "T00:00:00"))
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid date: {value} (expected YYYY-MM-DD)")
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def _date_iso_range_filter(field: str, date_from: Optional[str], date_to: Optional[str]) -> dict:
    """Build a Mongo filter for an ISO-string date field stored as 'YYYY-MM-DDTHH:MM:SS...' """
    if not date_from and not date_to:
        return {}
    rng: dict = {}
    if date_from:
        rng["$gte"] = date_from + "T00:00:00"
    if date_to:
        rng["$lte"] = date_to + "T23:59:59"
    return {field: rng}


def _date_str_range_filter(field: str, date_from: Optional[str], date_to: Optional[str]) -> dict:
    """For YYYY-MM-DD only fields (period_date, week_end_date)."""
    if not date_from and not date_to:
        return {}
    rng: dict = {}
    if date_from:
        rng["$gte"] = date_from
    if date_to:
        rng["$lte"] = date_to
    return {field: rng}


async def _resolve_user_ids_for_query(db, q: Optional[str]) -> Optional[List[str]]:
    """Convert a free-text query ('john', 'foo@bar.com', 'user_abc123') to a set of user_ids.
    Returns None when no filter was requested, [] when filter matched nothing."""
    if not q:
        return None
    q = q.strip()
    if not q:
        return None
    pattern = re.compile(re.escape(q), re.IGNORECASE)
    cursor = db.users.find(
        {
            "$or": [
                {"name": {"$regex": pattern}},
                {"email": {"$regex": pattern}},
                {"user_id": {"$regex": pattern}},
                {"referral_code": {"$regex": pattern}},
            ]
        },
        {"_id": 0, "user_id": 1},
    )
    ids = [u["user_id"] async for u in cursor]
    return ids


# ---------------------------------------------------------------------------
# Report builders — each returns (columns, rows, meta)
# ---------------------------------------------------------------------------
async def report_investments(db, *, date_from, date_to, q) -> Tuple[List[tuple], List[dict], dict]:
    flt: dict = {}
    flt.update(_date_iso_range_filter("created_at", date_from, date_to))
    user_ids = await _resolve_user_ids_for_query(db, q)
    if user_ids is not None:
        if not user_ids:
            return _investments_columns(), [], {"count": 0, "total_amount": 0}
        flt["user_id"] = {"$in": user_ids}

    # Build name map once
    name_map: Dict[str, str] = {}
    if user_ids:
        async for u in db.users.find({"user_id": {"$in": user_ids}}, {"_id": 0, "user_id": 1, "name": 1, "email": 1}):
            name_map[u["user_id"]] = f"{u.get('name') or '—'} ({u.get('email','')})"

    rows: List[dict] = []
    total = 0.0
    cursor = db.investments.find(flt, {"_id": 0}).sort("created_at", -1)
    pending_ids: List[str] = []
    async for inv in cursor:
        rows.append(inv)
        if inv.get("status") != "rejected":
            total += float(inv.get("amount", 0))
        if inv["user_id"] not in name_map:
            pending_ids.append(inv["user_id"])
    if pending_ids:
        async for u in db.users.find({"user_id": {"$in": list(set(pending_ids))}},
                                     {"_id": 0, "user_id": 1, "name": 1, "email": 1}):
            name_map[u["user_id"]] = f"{u.get('name') or '—'} ({u.get('email','')})"

    out_rows: List[dict] = []
    for inv in rows:
        out_rows.append({
            "created_at": (inv.get("created_at") or "")[:19].replace("T", " "),
            "user_id": inv["user_id"],
            "user": name_map.get(inv["user_id"], "—"),
            "plan_name": inv.get("plan_name"),
            "amount": float(inv.get("amount", 0)),
            "network": inv.get("network"),
            "tx_hash": inv.get("tx_hash") or "—",
            "mobile_number": inv.get("mobile_number") or "—",
            "status": inv.get("status"),
            "activated_at": (inv.get("activated_at") or "")[:19].replace("T", " "),
        })
    return _investments_columns(), out_rows, {"count": len(out_rows), "total_amount": round(total, 2)}


def _investments_columns():
    return [
        ("created_at", "Created"),
        ("user_id", "User ID"),
        ("user", "Investor"),
        ("plan_name", "Plan"),
        ("amount", "Amount (USDT)"),
        ("network", "Method"),
        ("tx_hash", "Tx Hash"),
        ("mobile_number", "Mobile"),
        ("status", "Status"),
        ("activated_at", "Activated"),
    ]


async def report_earnings(db, *, date_from, date_to, q) -> Tuple[List[tuple], List[dict], dict]:
    flt: dict = {}
    flt.update(_date_str_range_filter("period_date", date_from, date_to))
    user_ids = await _resolve_user_ids_for_query(db, q)
    if user_ids is not None:
        if not user_ids:
            return _earnings_columns(), [], {"count": 0, "total_amount": 0.0}
        flt["user_id"] = {"$in": user_ids}

    rows: List[dict] = []
    total = 0.0
    name_map: Dict[str, str] = {}
    async for e in db.earnings.find(flt, {"_id": 0}).sort([("period_date", -1), ("created_at", -1)]):
        rows.append(e)
        total += float(e.get("amount", 0))
    uids = list({r["user_id"] for r in rows})
    if uids:
        async for u in db.users.find({"user_id": {"$in": uids}}, {"_id": 0, "user_id": 1, "name": 1, "email": 1}):
            name_map[u["user_id"]] = f"{u.get('name') or '—'} ({u.get('email','')})"

    out_rows: List[dict] = []
    for e in rows:
        out_rows.append({
            "period_date": e.get("period_date"),
            "user_id": e["user_id"],
            "user": name_map.get(e["user_id"], "—"),
            "type": e.get("type"),
            "amount": float(e.get("amount", 0)),
            "source_level": e.get("source_level", 0),
            "status": e.get("status"),
            "payout_id": e.get("payout_id") or "—",
        })
    return _earnings_columns(), out_rows, {"count": len(out_rows), "total_amount": round(total, 6)}


def _earnings_columns():
    return [
        ("period_date", "Date"),
        ("user_id", "User ID"),
        ("user", "Investor"),
        ("type", "Type"),
        ("amount", "Amount (USDT)"),
        ("source_level", "Level"),
        ("status", "Status"),
        ("payout_id", "Payout"),
    ]


async def report_payouts(db, *, date_from, date_to, q) -> Tuple[List[tuple], List[dict], dict]:
    flt: dict = {}
    flt.update(_date_str_range_filter("week_end_date", date_from, date_to))
    user_ids = await _resolve_user_ids_for_query(db, q)
    if user_ids is not None:
        if not user_ids:
            return _payouts_columns(), [], {"count": 0, "total_amount": 0.0}
        flt["user_id"] = {"$in": user_ids}

    rows: List[dict] = []
    total = 0.0
    async for p in db.weekly_payouts.find(flt, {"_id": 0}).sort("week_end_date", -1):
        rows.append(p)
        total += float(p.get("total", 0))
    uids = list({r["user_id"] for r in rows})
    name_map: Dict[str, str] = {}
    if uids:
        async for u in db.users.find({"user_id": {"$in": uids}}, {"_id": 0, "user_id": 1, "name": 1, "email": 1}):
            name_map[u["user_id"]] = f"{u.get('name') or '—'} ({u.get('email','')})"

    out_rows: List[dict] = []
    for p in rows:
        bd = p.get("breakdown") or {}
        out_rows.append({
            "week_end_date": p.get("week_end_date"),
            "user_id": p["user_id"],
            "user": name_map.get(p["user_id"], "—"),
            "total": float(p.get("total", 0)),
            "roi": float(bd.get("roi", 0)),
            "level": float(bd.get("level", 0)),
            "referral": float(bd.get("referral", 0)),
            "earning_count": p.get("earning_count", 0),
            "status": p.get("status"),
        })
    return _payouts_columns(), out_rows, {"count": len(out_rows), "total_amount": round(total, 2)}


def _payouts_columns():
    return [
        ("week_end_date", "Week ending"),
        ("user_id", "User ID"),
        ("user", "Investor"),
        ("total", "Total (USDT)"),
        ("roi", "ROI"),
        ("level", "Level"),
        ("referral", "Referral"),
        ("earning_count", "# Earnings"),
        ("status", "Status"),
    ]


async def report_investors(db, *, date_from, date_to, q) -> Tuple[List[tuple], List[dict], dict]:
    flt: dict = {"role": "investor"}
    flt.update(_date_iso_range_filter("created_at", date_from, date_to))
    user_ids = await _resolve_user_ids_for_query(db, q)
    if user_ids is not None:
        if not user_ids:
            return _investors_columns(), [], {"count": 0, "total_active_capital": 0.0}
        flt["user_id"] = {"$in": user_ids}

    users = []
    async for u in db.users.find(flt, {"_id": 0, "password_hash": 0}).sort("created_at", -1):
        users.append(u)

    # Active capital + direct referrals + team size
    uids = [u["user_id"] for u in users]
    cap_map: Dict[str, float] = {}
    if uids:
        async for r in db.investments.aggregate([
            {"$match": {"user_id": {"$in": uids}, "status": "active"}},
            {"$group": {"_id": "$user_id", "total": {"$sum": "$amount"}}},
        ]):
            cap_map[r["_id"]] = float(r.get("total", 0))

    # Build downline counts
    direct_counts: Dict[str, int] = {}
    async for u in db.users.find({"referred_by": {"$in": uids}}, {"_id": 0, "referred_by": 1}):
        direct_counts[u["referred_by"]] = direct_counts.get(u["referred_by"], 0) + 1

    # Sponsor email lookup
    sponsor_ids = {u["referred_by"] for u in users if u.get("referred_by")}
    sp_map: Dict[str, str] = {}
    if sponsor_ids:
        async for s in db.users.find({"user_id": {"$in": list(sponsor_ids)}},
                                     {"_id": 0, "user_id": 1, "name": 1, "email": 1}):
            sp_map[s["user_id"]] = s.get("email") or s.get("name") or s["user_id"]

    out_rows: List[dict] = []
    total_active = 0.0
    for u in users:
        cap = cap_map.get(u["user_id"], 0.0)
        total_active += cap
        out_rows.append({
            "created_at": (u.get("created_at") or "")[:19].replace("T", " "),
            "user_id": u["user_id"],
            "name": u.get("name") or "—",
            "email": u.get("email"),
            "country": u.get("country") or "—",
            "phone": u.get("phone") or "—",
            "referral_code": u.get("referral_code") or "—",
            "sponsor": sp_map.get(u.get("referred_by") or "", "—"),
            "direct_referrals": direct_counts.get(u["user_id"], 0),
            "active_capital": round(cap, 2),
            "is_active": "Yes" if u.get("is_active", True) else "No",
        })
    return _investors_columns(), out_rows, {"count": len(out_rows), "total_active_capital": round(total_active, 2)}


def _investors_columns():
    return [
        ("created_at", "Joined"),
        ("user_id", "User ID"),
        ("name", "Name"),
        ("email", "Email"),
        ("country", "Country"),
        ("phone", "Phone"),
        ("referral_code", "Ref Code"),
        ("sponsor", "Sponsor"),
        ("direct_referrals", "Directs"),
        ("active_capital", "Active Capital"),
        ("is_active", "Active"),
    ]


async def report_leads(db, *, date_from, date_to, q) -> Tuple[List[tuple], List[dict], dict]:
    flt: dict = {}
    flt.update(_date_iso_range_filter("created_at", date_from, date_to))
    if q:
        pattern = re.compile(re.escape(q.strip()), re.IGNORECASE)
        flt["$or"] = [
            {"email": {"$regex": pattern}},
            {"note": {"$regex": pattern}},
            {"ticket_size": {"$regex": pattern}},
            {"lead_id": {"$regex": pattern}},
        ]
    rows: List[dict] = []
    async for lead in db.leads.find(flt, {"_id": 0}).sort("created_at", -1):
        rows.append({
            "created_at": (lead.get("created_at") or "")[:19].replace("T", " "),
            "lead_id": lead.get("lead_id"),
            "email": lead.get("email"),
            "ticket_size": lead.get("ticket_size"),
            "note": (lead.get("note") or "")[:200],
        })
    return _leads_columns(), rows, {"count": len(rows)}


def _leads_columns():
    return [
        ("created_at", "Submitted"),
        ("lead_id", "Lead ID"),
        ("email", "Email"),
        ("ticket_size", "Ticket size"),
        ("note", "Note"),
    ]


# ---------------------------------------------------------------------------
# Renderers
# ---------------------------------------------------------------------------
def _xlsx_response(columns, rows, meta, *, title, filename) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    # Excel doesn't allow / \ ? * [ ] : in sheet titles
    safe_title = re.sub(r"[\\/\?\*\[\]:]", " ", title)[:31].strip() or "Report"
    ws.title = safe_title

    # Title row
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
    # Meta row
    meta_str = " · ".join(f"{k}: {v}" for k, v in meta.items())
    ws["A2"] = meta_str + f"  · Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = Font(size=9, italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(columns), 1))

    # Headers
    header_fill = PatternFill(start_color="0F5E3F", end_color="0F5E3F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for ci, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=ci, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")

    # Data
    for ri, row in enumerate(rows, start=5):
        for ci, (k, _) in enumerate(columns, start=1):
            ws.cell(row=ri, column=ci, value=row.get(k))

    # Auto-width
    for ci, (k, label) in enumerate(columns, start=1):
        col_letter = get_column_letter(ci)
        max_len = len(str(label))
        for row in rows:
            v = row.get(k)
            if v is not None:
                vl = len(str(v))
                if vl > max_len:
                    max_len = vl
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


def _pdf_response(columns, rows, meta, *, title, filename) -> StreamingResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24,
        title=title,
    )
    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Heading1"], fontSize=14, leading=18, textColor=colors.HexColor("#0A0A0A"))
    m = ParagraphStyle("m", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#666666"))
    cell_p = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7.5, leading=9.5,
                            textColor=colors.HexColor("#0A0A0A"))

    story = [
        Paragraph(title, h),
        Paragraph(
            " · ".join(f"<b>{k}</b>: {v}" for k, v in meta.items())
            + f" · Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            m,
        ),
        Spacer(1, 8),
    ]

    if not rows:
        story.append(Paragraph("<i>No data for the selected filters.</i>", m))
    else:
        def _fmt(v):
            if v is None:
                return ""
            if isinstance(v, float):
                return f"{v:,.2f}"
            return str(v)

        header = [Paragraph(f"<b>{label}</b>", cell_p) for _, label in columns]
        data = [header]
        for row in rows:
            data.append([Paragraph(_fmt(row.get(k)), cell_p) for k, _ in columns])

        n_cols = len(columns)
        # Approximate column widths so the table fits the page
        page_w = landscape(A4)[0] - 48
        col_w = [page_w / n_cols] * n_cols

        tbl = Table(data, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F5E3F")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, 0), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
            ("TOPPADDING",    (0, 0), (-1, 0), 5),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FBFAF6")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FBFAF6")]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E1D6")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 1), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 2),
        ]))
        story.append(tbl)

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'},
    )


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
REPORTS: Dict[str, Tuple[Callable, str]] = {
    "investments": (report_investments, "Deposits / Investments report"),
    "earnings":    (report_earnings,    "Earnings ledger report"),
    "payouts":     (report_payouts,     "Weekly payouts report"),
    "investors":   (report_investors,   "Investors report"),
    "leads":       (report_leads,       "Allocator leads report"),
}


async def run_report(db, *, name: str, fmt: str, date_from: Optional[str],
                     date_to: Optional[str], q: Optional[str]):
    if name not in REPORTS:
        raise HTTPException(status_code=404, detail=f"Unknown report: {name}")
    fn, title = REPORTS[name]
    # Validate date strings
    _parse_dt(date_from)
    _parse_dt(date_to, end_of_day=True)

    columns, rows, meta = await fn(db, date_from=date_from, date_to=date_to, q=q)
    meta_for_header = {"Rows": meta.get("count", len(rows))}
    if "total_amount" in meta:
        meta_for_header["Total"] = f"{meta['total_amount']:,}"
    if "total_active_capital" in meta:
        meta_for_header["Active Capital"] = f"{meta['total_active_capital']:,}"
    if date_from:
        meta_for_header["From"] = date_from
    if date_to:
        meta_for_header["To"] = date_to
    if q:
        meta_for_header["Search"] = q

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    filename = f"{name}-report-{stamp}"

    if fmt == "json":
        return JSONResponse({
            "name": name,
            "title": title,
            "columns": [{"key": k, "label": label} for k, label in columns],
            "rows": rows,
            "meta": meta,
        })
    if fmt == "xlsx":
        return _xlsx_response(columns, rows, meta_for_header, title=title, filename=filename)
    if fmt == "pdf":
        return _pdf_response(columns, rows, meta_for_header, title=title, filename=filename)
    raise HTTPException(status_code=400, detail="format must be json, pdf, or xlsx")
